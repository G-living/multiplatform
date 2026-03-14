"""
Genera TTT_catalog_audit.xlsx desde ttt-catalog-data.js
Columnas: sku_producto | nombre_patron | categoria | zona | material |
          descripcion | fotos_producto | sku_variante | medida | precio | foto_variante
"""
import re, json, sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = Path(__file__).parent.parent / "js" / "ttt-catalog-data.js"
OUT = Path(__file__).parent / "TTT_catalog_audit.xlsx"

# ── extraer via Node.js (evita ambigüedades JS→JSON) ─────────────────────────
import subprocess, tempfile, os

loader = f"""
let src = require('fs').readFileSync('{SRC}','utf8');
src = src.replace("'use strict';","");
src = src.replace(/window\\.TTT_PRODUCTS\\s*=\\s*TTT_PRODUCTS;?/,'');
src = src.replace(/if\\s*\\(typeof Logger.*$/m,'');
src = src + '\\nmodule.exports = TTT_PRODUCTS;';
const tmp = require('os').tmpdir() + '/ttt_loader_gen.js';
require('fs').writeFileSync(tmp, src);
const data = require(tmp);
process.stdout.write(JSON.stringify(data));
"""
tmp_loader = "/tmp/ttt_gen_runner.js"
with open(tmp_loader, "w") as f:
    f.write(loader)

result = subprocess.run(["node", tmp_loader], capture_output=True, text=True)
if result.returncode != 0:
    sys.exit(f"Node error: {result.stderr}")

products = json.loads(result.stdout)

# ── estilos ───────────────────────────────────────────────────────────────────
GOLD   = "C9A96E"
DARK   = "1A1A1A"
LIGHT  = "FAF7F2"
ALT    = "F3EEE5"
WHITE  = "FFFFFF"

hdr_font   = Font(bold=True, color=WHITE, size=10, name="Calibri")
hdr_fill   = PatternFill("solid", fgColor=DARK)
gold_fill  = PatternFill("solid", fgColor=GOLD)
light_fill = PatternFill("solid", fgColor=LIGHT)
alt_fill   = PatternFill("solid", fgColor=ALT)
data_font  = Font(size=9, name="Calibri")
center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="D0C8BC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TTT Catálogo"
ws.freeze_panes = "A3"

HEADERS = [
    "SKU Producto",
    "Nombre / Patrón",
    "Categoría",
    "Zona",
    "Material",
    "Descripción",
    "Fotos del producto (todas)",
    "SKU Variante",
    "Medida",
    "Precio (€)",
    "Foto variante (nombre archivo)",
]

COL_W = [22, 18, 18, 16, 20, 30, 50, 24, 20, 10, 30]

# fila 1: título
ws.merge_cells("A1:K1")
title_cell = ws["A1"]
title_cell.value = "TTT — Auditoría catálogo catalog-data.js"
title_cell.font = Font(bold=True, color=WHITE, size=12, name="Calibri")
title_cell.fill = PatternFill("solid", fgColor=GOLD)
title_cell.alignment = center
ws.row_dimensions[1].height = 22

# fila 2: cabeceras
for col, (h, w) in enumerate(zip(HEADERS, COL_W), start=1):
    c = ws.cell(row=2, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = center
    c.border = border
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[2].height = 32

# ── datos ─────────────────────────────────────────────────────────────────────
row = 3
for idx, p in enumerate(products):
    fotos_str    = " | ".join(p.get("images", []))
    medidas      = p.get("medidas", [])
    fill_product = light_fill if idx % 2 == 0 else alt_fill

    for m_idx, m in enumerate(medidas):
        is_first = m_idx == 0
        fill_row = fill_product

        values = [
            p.get("sku", ""),
            p.get("name", "") + (" / " + p.get("patron","") if p.get("patron","") != p.get("name","") else ""),
            p.get("categoria", ""),
            p.get("zona", ""),
            p.get("material", ""),
            p.get("shortDesc", ""),
            fotos_str if is_first else "",
            m.get("sku", ""),
            m.get("medida", ""),
            m.get("precio", ""),
            m.get("imagen", ""),
        ]

        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = data_font
            c.fill = fill_row
            c.border = border
            c.alignment = left if col in (6, 7) else center

        ws.row_dimensions[row].height = 16
        row += 1

    # si no tiene medidas: fila con vacíos en columnas de variante
    if not medidas:
        values = [
            p.get("sku", ""),
            p.get("name", ""),
            p.get("categoria", ""),
            p.get("zona", ""),
            p.get("material", ""),
            p.get("shortDesc", ""),
            fotos_str,
            "⚠ SIN VARIANTES", "", "", "",
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(size=9, color="CC0000", name="Calibri")
            c.fill = PatternFill("solid", fgColor="FFF0F0")
            c.border = border
            c.alignment = center
        ws.row_dimensions[row].height = 16
        row += 1

# ── autofilter ────────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A2:K{row-1}"

# ── stats sheet ───────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Resumen")
from collections import Counter
cats  = Counter(p.get("categoria","") for p in products)
zonas = Counter(p.get("zona","")      for p in products)

ws2["A1"] = "Categoría";  ws2["B1"] = "# Productos"
ws2["D1"] = "Zona";       ws2["E1"] = "# Productos"
for c in ["A1","B1","D1","E1"]:
    ws2[c].font = Font(bold=True, color=WHITE, name="Calibri")
    ws2[c].fill = PatternFill("solid", fgColor=DARK)
    ws2[c].alignment = center

for i, (cat, n) in enumerate(sorted(cats.items()), start=2):
    ws2.cell(row=i, column=1, value=cat)
    ws2.cell(row=i, column=2, value=n)

for i, (zona, n) in enumerate(sorted(zonas.items()), start=2):
    ws2.cell(row=i, column=4, value=zona)
    ws2.cell(row=i, column=5, value=n)

for col in ["A","B","D","E"]:
    ws2.column_dimensions[col].width = 22

wb.save(OUT)
print(f"✓ Generado: {OUT}")
print(f"  {len(products)} productos · {row-3} filas de variantes")
