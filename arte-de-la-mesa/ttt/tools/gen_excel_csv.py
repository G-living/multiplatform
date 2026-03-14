"""
TTT — Excel de auditoría desde CSV fuente (Magento export)
Columnas: A=Familia | B=SKU variante | C=Nombre/Descripción | D=Precio EUR
          E=Precio COP (×18.250) | F=Material + Fotos (con label CSV)

Fuente: ttt_export.csv (export_catalog_product TTT)
Filtro: 6 familias arte-de-la-mesa
"""
import csv, re, sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_PATH = Path(__file__).parent / "ttt_export.csv"
OUT_PATH = Path(__file__).parent / "TTT_6familias_desde_CSV.xlsx"

COP_MULT = 18_250

FAMILY_MAP = {
    "TTT/B2C/tavola e cucina/tovaglie":   "Manteles (tovaglie)",
    "TTT/B2C/tavola e cucina/runner":     "Caminos de mesa (runner)",
    "TTT/B2C/tavola e cucina/tovaglioli": "Servilletas (tovaglioli)",
    "TTT/B2C/tavola e cucina/tovaglietta":"Individuales (tovaglietta)",
    "TTT/B2C/home decor/mezzero":         "Paños deco (mezzero)",
    "TTT/B2C/home decor/cuscini":         "Cojines (cuscini)",
}

# ── parser additional_attributes ─────────────────────────────────────────────
def parse_attrs(s):
    """Parsea key=value,key2=value2 — keys son letras/guion/dígito."""
    d = {}
    if not s:
        return d
    # lookahead: coma seguida de key= (letras, dígitos, guión bajo)
    parts = re.split(r',(?=[A-Za-z_][A-Za-z0-9_]*=)', s)
    for part in parts:
        if '=' in part:
            k, v = part.split('=', 1)
            d[k.strip()] = v.strip()
    return d

def get_images(attrs):
    """Retorna lista de (label_csv, filename) para las imágenes del producto."""
    image_keys = [
        "base_image",
        "small_image",
        "thumbnail_image",
        "ttt_image1",
        "ttt_image2",
        "ttt_image_horizontal",
        "ttt_theme_image",
        "ttt_gallery_images",
    ]
    result = []
    for key in image_keys:
        val = attrs.get(key, "")
        if val and val.lower() not in ("", "no_selection"):
            # ttt_gallery_images puede tener múltiples separadas por |
            if "|" in val:
                for i, img in enumerate(val.split("|")):
                    img = img.strip()
                    if img:
                        result.append((f"{key}[{i+1}]", img))
            else:
                result.append((key, val))
    return result

# ── cargar CSV ────────────────────────────────────────────────────────────────
grouped = {}   # sku -> row  (producto agrupado en las 6 familias)
simples = {}   # ttt_erp_name -> [row, row, ...]  (variantes simples)

with open(CSV_PATH, encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        # solo filas base (sin store_view_code)
        if row["store_view_code"] != "":
            continue
        cat = row["categories"]
        ptype = row["product_type"]

        if ptype == "grouped" and cat in FAMILY_MAP:
            grouped[row["sku"]] = row

        elif ptype == "simple":
            attrs = parse_attrs(row["additional_attributes"])
            erp = attrs.get("ttt_erp_name", "")
            if erp:
                simples.setdefault(erp, []).append((row, attrs))

print(f"Grouped 6 familias: {len(grouped)}")
total_variants = sum(len(v) for v in simples.values())
print(f"Simples totales: {total_variants}")

# ── estilos ───────────────────────────────────────────────────────────────────
GOLD  = "C9A96E"
DARK  = "1A1A1A"
LIGHT = "FAF7F2"
ALT   = "F3EEE5"
WHITE = "FFFFFF"
WARN  = "FFF0F0"
WARN_RED = "CC0000"

hdr_font  = Font(bold=True, color=WHITE, size=10, name="Calibri")
hdr_fill  = PatternFill("solid", fgColor=DARK)
gold_fill = PatternFill("solid", fgColor=GOLD)
data_font = Font(size=9, name="Calibri")
warn_font = Font(size=9, color=WARN_RED, name="Calibri")
center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin      = Side(style="thin", color="D0C8BC")
border    = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TTT 6 Familias"
ws.freeze_panes = "A3"

HEADERS = [
    "Familia",
    "SKU variante",
    "Nombre / Descripción",
    "Precio EUR",
    "Precio COP",
    "Material + Fotos (label CSV: archivo.jpg)",
]
COL_W = [26, 24, 36, 13, 16, 70]

# fila 1: título
ws.merge_cells("A1:F1")
t = ws["A1"]
t.value = "TTT — 6 Familias arte-de-la-mesa · Auditoría desde CSV fuente"
t.font   = Font(bold=True, color=WHITE, size=12, name="Calibri")
t.fill   = PatternFill("solid", fgColor=GOLD)
t.alignment = center
ws.row_dimensions[1].height = 22

# fila 2: cabeceras
for col, (h, w) in enumerate(zip(HEADERS, COL_W), start=1):
    c = ws.cell(row=2, column=col, value=h)
    c.font      = hdr_font
    c.fill      = hdr_fill
    c.alignment = center
    c.border    = border
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[2].height = 32

# ── datos ─────────────────────────────────────────────────────────────────────
row_num = 3
family_order = list(FAMILY_MAP.keys())

for g_idx, g_cat in enumerate(family_order):
    familia_label = FAMILY_MAP[g_cat]
    fill_base = PatternFill("solid", fgColor=LIGHT) if g_idx % 2 == 0 \
                else PatternFill("solid", fgColor=ALT)

    # productos de esta familia
    familia_prods = [(sku, r) for sku, r in grouped.items() if r["categories"] == g_cat]
    familia_prods.sort(key=lambda x: x[0])

    for p_sku, g_row in familia_prods:
        g_attrs   = parse_attrs(g_row["additional_attributes"])
        g_name    = g_row["name"]
        g_images  = get_images(g_attrs)   # fotos del grupo (nivel producto)

        variantes = simples.get(p_sku, [])

        if not variantes:
            # Sin variantes → fila de advertencia
            foto_str = "\n".join(f"{lbl}: {fn}" for lbl, fn in g_images) if g_images \
                       else "⚠ Sin fotos en CSV"
            vals = [familia_label, p_sku, g_name, "⚠ SIN VARIANTES", "", foto_str]
            for col, val in enumerate(vals, start=1):
                c = ws.cell(row=row_num, column=col, value=val)
                c.font      = warn_font
                c.fill      = PatternFill("solid", fgColor=WARN)
                c.border    = border
                c.alignment = left if col in (3, 6) else center
            ws.row_dimensions[row_num].height = max(16, 14 * max(1, len(g_images)))
            row_num += 1
            continue

        for v_row, v_attrs in variantes:
            price_eur = float(v_row["price"]) if v_row["price"] else 0.0
            price_cop = price_eur * COP_MULT

            # descripción: nombre patrón + medida variante
            v_name        = v_row["name"] or g_name
            measure_cm    = v_attrs.get("ttt_measure_short_descr_cm", "")
            measure_sub   = v_attrs.get("ttt_measure_subtype", "")
            typology      = v_attrs.get("ttt_typology_subtype", "")
            desc_parts    = [v_name]
            if measure_sub and measure_sub != v_name:
                desc_parts.append(measure_sub)
            if measure_cm and measure_cm != measure_sub:
                desc_parts.append(measure_cm)
            if typology and typology not in desc_parts:
                desc_parts.append(typology)
            description = " · ".join(desc_parts)

            # material + fotos
            material = v_attrs.get("ttt_material_desc", "")

            # fotos de la variante (pueden añadir ttt_image2, ttt_theme_image)
            v_images = get_images(v_attrs)

            # combinar: fotos del grupo primero, luego las de variante que no estén ya
            all_imgs = list(g_images)
            g_files  = {fn for _, fn in g_images}
            for lbl, fn in v_images:
                if fn not in g_files:
                    all_imgs.append((lbl, fn))

            foto_lines = [f"{lbl}: {fn}" for lbl, fn in all_imgs]
            if not foto_lines:
                foto_lines = ["⚠ Sin fotos en CSV"]

            col_f = (f"Material: {material}\n" if material else "") + \
                    "\n".join(foto_lines)

            vals = [
                familia_label,
                v_row["sku"],
                description,
                price_eur if price_eur else "⚠ Sin precio",
                int(price_cop) if price_cop else "",
                col_f,
            ]

            h = max(16, 14 * max(1, len(foto_lines) + (1 if material else 0)))

            for col, val in enumerate(vals, start=1):
                c = ws.cell(row=row_num, column=col, value=val)
                c.font      = data_font
                c.fill      = fill_base
                c.border    = border
                c.alignment = left if col in (3, 6) else center
                if col == 4 and isinstance(val, float):
                    c.number_format = '#,##0.00"€"'
                if col == 5 and isinstance(val, int):
                    c.number_format = '#,##0"COP"'

            ws.row_dimensions[row_num].height = h
            row_num += 1

# ── autofilter ────────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A2:F{row_num-1}"

# ── hoja resumen ──────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Resumen familias")

summary_headers = ["Familia", "# Variantes", "Con foto", "Sin foto", "Sin precio"]
for col, h in enumerate(summary_headers, start=1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color=WHITE, name="Calibri")
    c.fill = PatternFill("solid", fgColor=DARK)
    c.alignment = center
    ws2.column_dimensions[get_column_letter(col)].width = 26

for i, (g_cat, familia_label) in enumerate(FAMILY_MAP.items(), start=2):
    prods = [(sku, r) for sku, r in grouped.items() if r["categories"] == g_cat]
    total_v = 0; with_img = 0; no_img = 0; no_price = 0
    for p_sku, _ in prods:
        for v_row, v_attrs in simples.get(p_sku, []):
            total_v += 1
            all_img_keys = [k for k in v_attrs if "image" in k and v_attrs[k] not in ("","no_selection")]
            if all_img_keys:
                with_img += 1
            else:
                g_a = parse_attrs(grouped[p_sku]["additional_attributes"])
                g_imgs = get_images(g_a)
                if g_imgs:
                    with_img += 1
                else:
                    no_img += 1
            if not v_row["price"]:
                no_price += 1

    row_data = [familia_label, total_v, with_img, no_img, no_price]
    for col, val in enumerate(row_data, start=1):
        c = ws2.cell(row=i, column=col, value=val)
        c.alignment = center
        if col > 1 and isinstance(val, int) and val > 0 and col in (4, 5):
            c.font = Font(color=WARN_RED, name="Calibri")

wb.save(OUT_PATH)
print(f"\n✓ Generado: {OUT_PATH}")
print(f"  {row_num - 3} filas de variantes · {len(grouped)} productos agrupados")
